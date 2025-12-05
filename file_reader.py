import openpyxl
from openpyxl import Workbook, load_workbook, worksheet
from templates import (
    COMPANY_INFO_CELLS, 
    COMPANY_CELL_TITLES, 
    COMPANY_INFO_LIST, 
    COMPANY_CLOSE_CELLS
)

COMPANY_NAME_CELL = '1'

def read_file(ws: worksheet, company_dict: dict) -> None:
    """
    Reads the predetermined file (Day 1, Day 2, or Day 3) and populates the company dictionary with a single line
    worth of information to simulate the current moment in time.

    @param ws the current opened excel worksheet.
    @param company_dict the dictionary that will be updated with each call of this function.
    """

    for idx, cell in enumerate(COMPANY_CELL_TITLES):
        for index, row in enumerate(COMPANY_INFO_CELLS[idx]):
            company_dict[ws[f'{cell}{COMPANY_NAME_CELL}'].value][COMPANY_INFO_LIST[index]] = ws[
                f'{row}{company_dict["Start Cell"]}'
            ].value
    
    company_dict['Start Cell'] = str(int(company_dict['Start Cell']) + 1)

    return None

def read_bids_for_percentages(ws: worksheet, company_list: list) -> dict:
    """
    Reads the first bid for each company as well as the first bid at 4pm.

    @param ws the active excel worksheet.
    @param company_dict the dictionary containing all company info.

    @return a dictionary containing each company as keys and the first bid and close bid as keys and values
    within.
    """

    bid_dict = {}

    # Get first bid and build dictionary
    for company in company_list:
        bid_dict[company] = {
            "First Bid": ws[f'{COMPANY_CELL_TITLES[company_list.index(company)]}3'].value,
            "Close Price": 0.0
        }

    # Fill out close price with first instance found at 4pm signaling the beginning of after hours
    for index, row in enumerate(ws):
        if type(row[0]) == openpyxl.cell.cell.MergedCell:
            continue

        if row[0].value[1] == '4' and bid_dict[company_list[0]]['Close Price'] == 0.0:
            for company in company_list:
                bid_dict[company]["Close Price"] = ws[
                    f'{COMPANY_CLOSE_CELLS[company_list.index(company)]}{index}'
                ].value
                
            break

    return bid_dict

def read_time_from_file(ws: worksheet, cell: str) -> str:
    """
    Reads the timestamp from the file associated with the information line.

    @param ws the excel worksheet.
    @param cell the string containing the row number.

    @return the string of the timestamp.
    """

    return ws[f'A{cell}'].value

def open_workbook(path: str) -> worksheet:
    """
    Opens the excel workbook to be utilized throughout the script.

    @param path the string path to the desired excel file.

    @return an open and active worksheet to be read from.
    """
    wb = load_workbook(path)
    ws = wb.active

    return ws

def close_workbook(wb: Workbook) -> None:
    """
    Closing the active excel Workbook.

    @param wb the active excel Workbook.
    """
    wb.close()

    return None

### For Testing Purposes Only ###
# if __name__ == '__main__':
#     company_dict = {}

#     day_two_companies_list = DAY_TWO_COMPANIES.split(',')

#     for company in day_two_companies_list:
#         company_dict[company] = copy.deepcopy(COMPANY_INFO_TEMPLATE)

#     company_dict["Start Cell"] = INFO_START_CELL
    # DAY_TWO_PATHING = "Days/Day2.xlsx"
    # ws = open_workbook(DAY_TWO_PATHING)

    # for row in ws:
    #     print(f'ROW: {row[0].value}')
    #     break

#     read_file(ws, company_dict)

#     for c in company_dict:
#         print(c)
#         print(company_dict[c])
